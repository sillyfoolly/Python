import time
import openpyxl
from typing import List
from random import randint


def linear_search(element: int, arr: List):
    result = []
    for i in range(len(arr)):
        if arr[i] == element:
            result.append(i)
    if len(result) == 0:
        return element, ['Такого элемента нет']
    return element, result


def binary_search(element: int, arr: List):
    left = 0
    right = len(arr) - 1
    index = -1
    while left <= right:
        mid = (left + right) // 2
        if arr[mid] == element:
            index = mid
            right = mid - 1
        elif element < arr[mid]:
            right = mid - 1
        else:
            left = mid + 1
    if index == -1:
        return element, 'Такого элемента нет'
    return element, index


def xl_writer():
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(1, 1).value = 'Линейный поиск:'
    sheet.cell(2, 1).value = 'Бинарный поиск:'

    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 37
    sheet.column_dimensions['D'].width = 5
    sheet.column_dimensions['E'].width = 36
    sheet.column_dimensions['F'].width = 5

    left = time.perf_counter()
    res1 = [linear_search(randint(5, 33), x) for _ in range(5)]
    sheet.cell(1, 2).value = f'{(time.perf_counter() - left) * 1000} мс'

    left = time.perf_counter()
    res2 = [binary_search(randint(5, 33), sorted(x)) for _ in range(5)]
    sheet.cell(2, 2).value = f'{(time.perf_counter() - left) * 1000} мс'

    sheet.merge_cells('C1:D1')
    sheet.cell(1, 3).value = 'Результат линейного поиска:'
    last = 2
    for i in range(len(res1)):
        if type(res1[i][1][0]) != str:
            sheet.cell(last, 3).value = f'Элемент {res1[i][0]} находится под индексами:'
            for j in range(len(res1[i][1])):
                sheet.cell(last + j, 4).value = res1[i][1][j]
        else:
            sheet.cell(last, 3).value = f'Элемента {res1[i][0]} нет в списке'
        last += len(res1[i][1])

    sheet.merge_cells('E1:F1')
    sheet.cell(1, 5).value = 'Результат бинарного поиска:'
    for i in range(len(res2)):
        sheet.cell(i + 2, 5).value = f'Элемент {res2[i][0]} находится под индексом:'
        sheet.cell(i + 2, 6).value = res2[i][1]

    wb.save('res2.xlsx')


if __name__ == '__main__':
    x = [19,  23,  25,  33,  26,  10,  31,  33,  12,  18,  9,  7,  28,  29,  18,  33,  18,  9,  9,  24,  15,  31,  17,  5,  7,  27,  26,  14,  15,  23,  24,  6,  20,  31,  19,  27,  18,  30,  29,  7,  31,  9,  24,  9,  5,  32,  33,  31,  6,  18,  24,  22,  21,  23,  10, 2, 8,  5,  29,  26,  32,  8,  27,  31,  7]

    xl_writer()
