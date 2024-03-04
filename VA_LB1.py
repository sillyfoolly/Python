# random_list_of_nums = [19, 23, 25, 33, 26, 10, 31, 33, 12, 18, 9, 7, 28, 29, 18, 33, 18, 9, 9, 24, 15, 31, 17, 5, 7, 27, 26, 14, 15, 23, 24, 6, 20, 31, 19, 27, 18, 30, 29, 7, 31, 9, 24, 9, 5, 32, 33, 31, 6, 18, 24, 22, 21, 23, 10, 2, 8, 5, 29, 26, 32, 8, 27, 31, 7]
import time
import openpyxl

def bubble(arr):
    for i in range(len(arr) - 1):
        for j in range(len(arr) - i - 1):
            if arr[j] > arr[j + 1]:
                arr[j], arr[j + 1] = arr[j + 1], arr[j]

def merge_sort(nums):
    if len(nums) > 1:
        mid = len(nums) // 2
        left = nums[:mid]
        right = nums[mid:]
        merge_sort(left)
        merge_sort(right)
        i = j = k = 0
        while i < len(left) and j < len(right):
            if left[i] < right[j]:
                nums[k] = left[i]
                i += 1
            else:
                nums[k] = right[j]
                j += 1
            k += 1
        while i < len(left):
            nums[k] = left[i]
            i += 1
            k += 1
        while j < len(right):
            nums[k] = right[j]
            j += 1
            k += 1

def xl_writer():
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.cell(1, 1).value = 'Сортировка пузырьком:'
    sheet.cell(2, 1).value = 'Сортировка слиянием:'
    sheet.cell(3, 1).value = 'Стандартная сортировка:'

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30

    start = time.perf_counter()
    bubble(x1)
    sheet.cell(1, 2).value = f'{(time.perf_counter() - start) * 1000} мс'

    start = time.perf_counter()
    merge_sort(x2)
    sheet.cell(2, 2).value = f'{(time.perf_counter() - start) * 1000} мс'

    start = time.perf_counter()
    x3.sort()
    sheet.cell(3, 2).value = f'{(time.perf_counter() - start) * 1000} мс'

    sheet.cell(1, 3).value = 'Результат сортировки:'

    for i in range(len(x1)):
        sheet.cell(i + 1, 4).value = x1[i]

    wb.save('res.xlsx')

if __name__ == '__main__':
    x1 = x2 = x3 = [19, 23, 25, 33, 26, 10, 31, 33, 12, 18, 9, 7, 28, 29, 18, 33, 18, 9, 9, 24, 15, 31, 17, 5, 7, 27, 26, 14, 15, 23, 24, 6, 20, 31, 19, 27, 18, 30, 29, 7, 31, 9, 24, 9, 5, 32, 33, 31, 6, 18, 24, 22, 21, 23, 10, 2, 8, 5, 29, 26, 32, 8, 27, 31, 7]

    xl_writer()
